from openpyxl import load_workbook
from os.path import *


class Areas:
    def __init__(self): # load_workbook 경로 바꿔서 사용해야 돌아감. 통합위치정보.xlsx 경로 본인 컴퓨터 경로로 변경할 것
        self.filepath = abspath('xls').replace('server', '') + "/통합위치정보.xlsx"
        self.areas = {}
        self.load_wb = load_workbook(filename=self.filepath, data_only=True)
        self.load_ws = self.load_wb['Sheet1']

        for i in range(2, 250):
            self.areas[(self.load_ws['A' + str(i)].value + ' ' + self.load_ws['B' + str(i)].value)] = \
                ([(str(self.load_ws['D' + str(i)].value), str(self.load_ws['E' + str(i)].value)),
                  self.load_ws['F' + str(i)].value, self.load_ws['G' + str(i)].value, self.load_ws['H' + str(i)].value])


    def showCodes(self):
        print(self.areas)

    def getAreaList(self):
        # 지역명을 리스트로 반환해줌
        return self.areas.keys()


def Test():
    area = Areas()
    area.getCodesfromXl()
    area.showCodes()

